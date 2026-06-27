"""
generate_reports.py – Build the Excel dashboard and Markdown reports.
"""

import logging
from datetime import date

import pandas as pd

from src.config import (
    DASHBOARD_XLSX, DAILY_REPORT_MD, STRATEGIC_REPORT_MD,
    STRATEGIC_MARKETS, PERSONAS,
)

logger = logging.getLogger(__name__)

# ─── Excel Dashboard ─────────────────────────────────────────────────────────

def generate_excel_dashboard(
    df: pd.DataFrame,
    heatmaps: dict,
) -> None:
    """Write a multi-sheet Excel workbook to DASHBOARD_XLSX."""
    DASHBOARD_XLSX.parent.mkdir(parents=True, exist_ok=True)

    top_priority = (
        df.sort_values("priority_score", ascending=False)
        .head(50)[
            [
                "full_name", "company_clean", "position_clean",
                "persona", "area", "seniority", "strategic_market",
                "priority_score", "recommended_action", "connected_on_clean",
                "url",
            ]
        ]
        .reset_index(drop=True)
    )

    # Summary sheet data
    total = len(df)
    summary_rows = [
        ["LinkedIn Network Analysis", ""],
        ["Generated On", str(date.today())],
        ["Total Connections Analyzed", total],
        ["", ""],
        ["─── Persona Distribution ───", ""],
    ]
    for persona, cnt in df["persona"].value_counts().head(15).items():
        summary_rows.append([persona, cnt])

    summary_rows += [
        ["", ""],
        ["─── Market Distribution ───", ""],
    ]
    for mkt, cnt in df["strategic_market"].value_counts().items():
        pct = round(cnt / total * 100, 1)
        summary_rows.append([mkt, f"{cnt} ({pct}%)"])

    summary_rows += [
        ["", ""],
        ["─── High Priority (score ≥ 70) ───", ""],
        ["Count", int((df["priority_score"] >= 70).sum())],
        ["─── Medium Priority (40–69) ───", ""],
        ["Count", int(((df["priority_score"] >= 40) & (df["priority_score"] < 70)).sum())],
    ]

    summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    # Classified connections (selected columns)
    classified_export = df[[
        "full_name", "company_clean", "position_clean", "connected_on_clean",
        "persona", "area", "seniority", "strategic_market",
        "market_confidence", "inference_reason",
        "priority_score", "recommended_action", "url",
    ]].copy()

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(DASHBOARD_XLSX, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            classified_export.to_excel(writer, sheet_name="Classified Connections", index=False)
            heatmaps["persona"].to_excel(writer, sheet_name="Persona Heatmap", index=False)
            heatmaps["area"].to_excel(writer, sheet_name="Area Heatmap", index=False)
            heatmaps["seniority"].to_excel(writer, sheet_name="Seniority Heatmap", index=False)
            heatmaps["market"].to_excel(writer, sheet_name="Market Heatmap", index=False)
            heatmaps["company"].to_excel(writer, sheet_name="Company Heatmap", index=False)
            heatmaps["gap"].to_excel(writer, sheet_name="Strategic Gap", index=False)
            top_priority.to_excel(writer, sheet_name="Top Priority Contacts", index=False)

            # Auto-fit column widths for each sheet
            wb = writer.book
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for col_cells in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col_cells[0].column)
                    for cell in col_cells:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

                # Style header row
                header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

        logger.info(f"Excel dashboard saved: {DASHBOARD_XLSX}")

    except ImportError:
        logger.warning("openpyxl not found – writing plain Excel via xlsxwriter.")
        with pd.ExcelWriter(DASHBOARD_XLSX, engine="xlsxwriter") as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            classified_export.to_excel(writer, sheet_name="Classified Connections", index=False)
            heatmaps["persona"].to_excel(writer, sheet_name="Persona Heatmap", index=False)
            heatmaps["area"].to_excel(writer, sheet_name="Area Heatmap", index=False)
            heatmaps["seniority"].to_excel(writer, sheet_name="Seniority Heatmap", index=False)
            heatmaps["market"].to_excel(writer, sheet_name="Market Heatmap", index=False)
            heatmaps["company"].to_excel(writer, sheet_name="Company Heatmap", index=False)
            heatmaps["gap"].to_excel(writer, sheet_name="Strategic Gap", index=False)
            top_priority.to_excel(writer, sheet_name="Top Priority Contacts", index=False)
        logger.info(f"Excel dashboard saved: {DASHBOARD_XLSX}")


# ─── Daily Markdown Report ────────────────────────────────────────────────────

def generate_daily_report(df: pd.DataFrame, heatmaps: dict) -> None:
    """Write the daily_network_report.md Markdown report."""
    DAILY_REPORT_MD.parent.mkdir(parents=True, exist_ok=True)

    total = len(df)
    today = date.today().strftime("%B %d, %Y")

    top_personas   = df["persona"].value_counts().head(10)
    top_areas      = df["area"].value_counts().head(8)
    top_seniority  = df["seniority"].value_counts().head(8)
    top_companies  = df["company_clean"].value_counts().head(10)
    market_dist    = df["strategic_market"].value_counts()

    high_pct = round((df["priority_score"] >= 70).sum() / total * 100, 1)
    med_pct  = round(((df["priority_score"] >= 40) & (df["priority_score"] < 70)).sum() / total * 100, 1)

    gap_df = heatmaps["gap"]
    critical_gaps = gap_df[gap_df["priority"] == "CRITICAL"].head(10)
    top_50 = (
        df.sort_values("priority_score", ascending=False)
        .head(50)[["full_name", "company_clean", "position_clean", "persona",
                   "strategic_market", "priority_score", "recommended_action"]]
        .reset_index(drop=True)
    )

    # Strengths & weaknesses
    strong_personas  = [p for p, c in top_personas.items() if c >= 20]
    weak_markets     = [m for m in STRATEGIC_MARKETS
                        if market_dist.get(m, 0) < 20 and m != "UNKNOWN"]
    strong_markets   = [m for m in STRATEGIC_MARKETS
                        if market_dist.get(m, 0) >= 50]

    lines = [
        f"# 📊 Daily LinkedIn Network Report",
        f"> Generated: **{today}**",
        "",
        "---",
        "",
        "## 📌 Executive Summary",
        "",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Total Connections Analyzed | **{total:,}** |",
        f"| High Priority (score ≥ 70) | **{(df['priority_score'] >= 70).sum():,}** ({high_pct}%) |",
        f"| Medium Priority (40–69) | **{((df['priority_score'] >= 40) & (df['priority_score'] < 70)).sum():,}** ({med_pct}%) |",
        f"| Report Date | {today} |",
        "",
        "---",
        "",
        "## 👥 Top Personas in Your Network",
        "",
        "| Persona | Count | % of Network |",
        "|---------|-------|--------------|",
    ]
    for persona, cnt in top_personas.items():
        lines.append(f"| {persona} | {cnt:,} | {round(cnt/total*100,1)}% |")

    lines += [
        "",
        "---",
        "",
        "## 🏢 Top Functional Areas",
        "",
        "| Area | Count | % of Network |",
        "|------|-------|--------------|",
    ]
    for area, cnt in top_areas.items():
        lines.append(f"| {area} | {cnt:,} | {round(cnt/total*100,1)}% |")

    lines += [
        "",
        "---",
        "",
        "## 🎯 Seniority Distribution",
        "",
        "| Seniority | Count |",
        "|-----------|-------|",
    ]
    for seniority, cnt in top_seniority.items():
        lines.append(f"| {seniority} | {cnt:,} |")

    lines += [
        "",
        "---",
        "",
        "## 🏭 Top Companies",
        "",
        "| Company | Connections |",
        "|---------|-------------|",
    ]
    for company, cnt in top_companies.items():
        if company.strip():
            lines.append(f"| {company} | {cnt} |")

    lines += [
        "",
        "---",
        "",
        "## 🌍 Strategic Market Distribution",
        "",
        "| Market | Count | % | Avg Priority Score |",
        "|--------|-------|---|--------------------|",
    ]
    mkt_df = heatmaps["market"]
    for _, row in mkt_df.iterrows():
        pct = round(row["count"] / total * 100, 1)
        lines.append(f"| {row['strategic_market']} | {int(row['count']):,} | {pct}% | {row['avg_priority']} |")

    lines += [
        "",
        "---",
        "",
        "## 💪 Network Strengths",
        "",
    ]
    if strong_personas:
        lines.append(f"- **Strong persona coverage:** {', '.join(strong_personas)}")
    if strong_markets:
        lines.append(f"- **Good market presence:** {', '.join(strong_markets)}")
    brazil_cnt = market_dist.get("BRAZIL", 0)
    lines.append(f"- **Brazil base:** {brazil_cnt:,} connections — solid local foundation.")
    recruiter_cnt = (df["persona"] == "Recruiter").sum() + (df["persona"] == "Talent Acquisition").sum()
    lines.append(f"- **Recruiter/TA network:** {recruiter_cnt:,} contacts — key pipeline for job leads.")

    lines += [
        "",
        "---",
        "",
        "## ⚠️ Network Weaknesses & Gaps",
        "",
    ]
    if weak_markets:
        lines.append(f"- **Underrepresented markets:** {', '.join(weak_markets)}")
    us_nearshore_cnt = market_dist.get("US_CANADA_NEARSHORE", 0)
    latam_usd_cnt    = market_dist.get("LATAM_USD", 0)
    spain_eu_cnt     = market_dist.get("SPAIN_EU", 0)
    lines += [
        f"- **US/Canada Nearshore:** Only {us_nearshore_cnt:,} connections — needs aggressive growth.",
        f"- **LATAM USD:** {latam_usd_cnt:,} connections — key for remote USD work from Brazil.",
        f"- **Spain/EU:** {spain_eu_cnt:,} connections — low for medium-term Europe strategy.",
        "- **Data Engineering Managers & Heads of Data** outside Brazil are underrepresented.",
    ]

    lines += [
        "",
        "---",
        "",
        "## 🚀 Short-Term Recommendations (Brazil → USD Remote Job)",
        "",
        "1. **Target US/Canada Nearshore recruiters & hiring managers** — send personalized outreach.",
        "2. **Engage LATAM USD staffing firms** (AgileEngine, Andela, Toptal, Wizeline, etc.).",
        "3. **Connect with Data Engineering Managers** at companies that hire remote LATAM contractors.",
        "4. **Reactivate top 50 high-priority contacts** from this report — send a short, focused message.",
        "5. **Post content on LinkedIn** about your Data Engineering skills to attract inbound recruiter attention.",
        "6. **Follow key companies**: Extractta, Indicium AI, Nimble.LA, AgileEngine, Blue People.",
        "",
        "---",
        "",
        "## 🌍 Medium-Term Recommendations (Spain / Europe Move)",
        "",
        "1. **Start building Spain/EU recruiter network now** — target Madrid, Barcelona, Dublin tech firms.",
        "2. **Connect with European Data Leaders** at ERNI, Adesso, NTT Data, Capgemini Europe, Deloitte.",
        "3. **Follow companies in Germany, Netherlands, Ireland** active in data/cloud roles.",
        "4. **Engage Portugal tech community** (close cultural/language bridge from Brazil).",
        "5. **Join Spain/EU expat & tech groups** on LinkedIn to increase inbound connections.",
        "",
        "---",
        "",
        "## 🎯 Top 10 Priority Connection Categories to Build",
        "",
    ]
    top_gaps = gap_df.head(10)
    for i, row in top_gaps.iterrows():
        lines.append(
            f"{i+1}. **{row['market']} – {row['persona']}**: "
            f"Need {row['short_term_gap']} more (current: {row['current_count']}, "
            f"target: {row['short_term_target']}) [{row['priority']}]"
        )

    lines += [
        "",
        "---",
        "",
        "## 🏆 Top 50 High-Priority Existing Connections to Review",
        "",
        "| # | Name | Company | Position | Persona | Market | Score | Action |",
        "|---|------|---------|----------|---------|--------|-------|--------|",
    ]
    for i, row in top_50.iterrows():
        lines.append(
            f"| {i+1} | {row['full_name']} | {row['company_clean']} | "
            f"{row['position_clean'][:40]}... | {row['persona']} | "
            f"{row['strategic_market']} | **{row['priority_score']}** | "
            f"{row['recommended_action'][:60]}... |"
        )

    lines += [
        "",
        "---",
        "",
        "_This report is auto-generated by the LinkedIn Connections Heatmap project._",
        "_See `outputs/` for CSV heatmaps and `reports/dashboard_ready.xlsx` for the full Excel dashboard._",
        "",
    ]

    DAILY_REPORT_MD.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"Daily report saved: {DAILY_REPORT_MD}")


# ─── Strategic Gap Markdown Report ───────────────────────────────────────────

def generate_strategic_gap_markdown(heatmaps: dict) -> None:
    """Write a dedicated strategic_gap_report.md."""
    STRATEGIC_REPORT_MD.parent.mkdir(parents=True, exist_ok=True)
    gap_df = heatmaps["gap"]
    today  = date.today().strftime("%B %d, %Y")

    lines = [
        "# 🎯 Strategic Gap Report",
        f"> Generated: **{today}**",
        "",
        "This report compares your current LinkedIn network against your strategic targets.",
        "",
        "## Interpretation",
        "- **short_term_gap**: connections needed for your Brazil → USD remote job goal.",
        "- **medium_term_gap**: connections needed for your Spain / Europe move.",
        "- **Priority CRITICAL**: Gap > 80% of target — take action immediately.",
        "",
        "---",
        "",
        "| Market | Persona | Current | ST Target | ST Gap | MT Target | MT Gap | Priority |",
        "|--------|---------|---------|-----------|--------|-----------|--------|----------|",
    ]
    for _, row in gap_df.iterrows():
        lines.append(
            f"| {row['market']} | {row['persona']} | {row['current_count']} | "
            f"{row['short_term_target']} | **{row['short_term_gap']}** | "
            f"{row['medium_term_target']} | {row['medium_term_gap']} | "
            f"**{row['priority']}** |"
        )

    lines += [
        "",
        "---",
        "",
        "## How to Use This Report",
        "",
        "1. Focus on **CRITICAL** rows first — these are your biggest network gaps.",
        "2. Use LinkedIn search to find professionals in these market/persona combinations.",
        "3. Send personalized connection requests explaining your goals.",
        "4. Re-run this report weekly to track progress.",
        "",
        "_Generated by LinkedIn Connections Heatmap — see README for instructions._",
    ]

    STRATEGIC_REPORT_MD.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"Strategic gap report saved: {STRATEGIC_REPORT_MD}")
