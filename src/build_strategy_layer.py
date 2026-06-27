# -*- coding: utf-8 -*-
"""
build_strategy_layer.py – Entry point for the strategic intelligence layer.

Reads the existing classified_connections.csv (produced by build_network_heatmap.py),
computes KPIs, action plans, gap matrices, dashboard JSON, and generates
strategic Markdown reports + updated Excel workbook.

Usage:
    python src/build_strategy_layer.py
"""

import io
import json
import logging
import sys
import time
from pathlib import Path

# ── UTF-8 stdout on Windows ─────────────────────────────────────────────────
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd

from src.config import (
    CLASSIFIED_CSV, OUTPUTS_DIR, REPORTS_DIR,
    DASHBOARD_XLSX,
)
from src.calculate_kpis import compute_kpis
from src.generate_action_plan import (
    build_30_day_plan, build_60_day_plan, build_90_day_plan,
    build_market_strategy_matrix, build_persona_strategy_matrix,
    build_connection_gap_matrix,
)
from src.generate_dashboard_data import (
    build_company_intelligence, build_heatmap_data,
    build_top_contacts, save_dashboard_json,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

# ── output paths ─────────────────────────────────────────────────────────────
KPI_CSV          = OUTPUTS_DIR / "kpi_summary.csv"
MARKET_MATRIX    = OUTPUTS_DIR / "market_strategy_matrix.csv"
PERSONA_MATRIX   = OUTPUTS_DIR / "persona_strategy_matrix.csv"
GAP_MATRIX       = OUTPUTS_DIR / "connection_gap_matrix.csv"
PLAN_30          = OUTPUTS_DIR / "action_plan_30_days.csv"
PLAN_60          = OUTPUTS_DIR / "action_plan_60_days.csv"
PLAN_90          = OUTPUTS_DIR / "action_plan_90_days.csv"
EXEC_REPORT      = REPORTS_DIR / "executive_strategy_report.md"
CAREER_ROADMAP   = REPORTS_DIR / "career_network_roadmap.md"
KPI_REPORT       = REPORTS_DIR / "kpi_dashboard_report.md"


def _save_csv(df: pd.DataFrame, path: Path, label: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    logger.info(f"Saved {label}: {path.name}")


def _save_kpi_csv(kpis: dict) -> None:
    rows = [{"kpi": k, "value": v} for k, v in kpis.items()
            if not isinstance(v, (dict, list))]
    df = pd.DataFrame(rows)
    _save_csv(df, KPI_CSV, "KPI Summary")


# ── Markdown report generators ───────────────────────────────────────────────

def _executive_report(kpis: dict, gap_df: pd.DataFrame) -> str:
    from datetime import date
    today = date.today().strftime("%B %d, %Y")
    total     = kpis["total_connections"]
    high      = kpis["high_priority"]
    usd_score = kpis["usd_opportunity_score"]
    spain_score = kpis["spain_readiness_score"]
    latam     = kpis["latam_usd_count"]
    us_near   = kpis["us_nearshore_count"]
    spain_eu  = kpis["spain_eu_count"]
    europe    = kpis["europe_count"]
    unknown   = kpis["unknown_count"]
    unknown_pct = kpis["unknown_pct"]
    recruiters  = kpis["recruiters_total"]
    ta          = kpis["talent_hr_total"]
    hm          = kpis["hiring_managers_total"]
    dl          = kpis["data_leaders_total"]
    flags       = kpis.get("concentration_flags", [])

    critical_gaps = gap_df[gap_df["urgency_level"] == "Critical"].head(5)

    usd_signal   = "YES" if usd_score >= 40 else "NOT YET"
    spain_signal = "EARLY STAGE" if spain_score >= 20 else "NOT STARTED"

    lines = [
        f"# Executive Strategy Report",
        f"> Generated: **{today}**",
        "",
        "---",
        "",
        "## 1. What Does My Current LinkedIn Network Look Like?",
        "",
        f"You have **{total:,} LinkedIn connections**, a large network by any standard.",
        f"However, quantity does not equal quality — only **{high:,} connections ({kpis['high_priority_pct']}%)**",
        "score above 70 (high strategic priority) for your current career objectives.",
        "",
        f"| Segment | Count |",
        f"|---------|-------|",
        f"| Total Connections | {total:,} |",
        f"| High Priority (>=70) | {high:,} |",
        f"| Recruiters | {recruiters:,} |",
        f"| Talent Acquisition + HR | {ta:,} |",
        f"| Hiring Managers | {hm:,} |",
        f"| Data Leaders | {dl:,} |",
        f"| LATAM USD market | {latam:,} |",
        f"| US/Canada Nearshore | {us_near:,} |",
        f"| Spain/EU | {spain_eu:,} |",
        f"| Europe | {europe:,} |",
        f"| Market UNKNOWN | {unknown:,} ({unknown_pct}%) |",
        "",
        "**Key insight:** The high UNKNOWN rate ({unknown_pct}%) is a data limitation, not a true gap.".format(unknown_pct=unknown_pct),
        "LinkedIn does not export location data. Market inference runs on company/title keywords only.",
        "",
        "---",
        "",
        "## 2. Is My Network Aligned With My Short-Term USD Goal?",
        "",
        f"**USD Opportunity Score: {usd_score}/100 — {usd_signal}**",
        "",
        f"- LATAM USD recruiters: {kpis['usd_recruiters']} | Target: 60",
        f"- US/CA Nearshore recruiters: {kpis['usd_ta']} | Target: 40",
        f"- USD Hiring Managers: {kpis['usd_hiring_mgrs']} | Target: 30",
        f"- USD Data Leaders: {kpis['usd_data_leaders']} | Target: 20",
        f"- High-priority USD contacts: {kpis['usd_high_priority']}",
        "",
        ("Your USD network is **developing but not mature**. "
         "You have a solid recruiter base in LATAM but need to aggressively grow "
         "US/Canada nearshore contacts and hiring manager relationships."),
        "",
        "---",
        "",
        "## 3. Is My Network Aligned With My Spain/Europe Plan?",
        "",
        f"**Spain Readiness Score: {spain_score}/100 — {spain_signal}**",
        "",
        f"- Spain/EU recruiters: {kpis['spain_recruiters']} | Target: 40",
        f"- Spain/EU Talent Acquisition: {kpis['spain_ta']} | Target: 30",
        f"- Spain/EU Hiring Managers: {kpis['spain_hiring_mgrs']} | Target: 20",
        f"- EU Data Leaders: {kpis['spain_data_leaders']} | Target: 15",
        f"- High-priority Spain/EU contacts: {kpis['spain_high_priority']}",
        "",
        ("Your Spain/EU network is **nascent**. "
         "You have some early contacts (ERNI, Stratesys, Deutsche Telekom) but far fewer than needed "
         "for a confident Europe job search. This is acceptable for now given your 6-18 month horizon."),
        "",
        "---",
        "",
        "## 4. What Are the Biggest Gaps?",
        "",
        "| Market | Persona | Current | Target | Gap | Urgency |",
        "|--------|---------|---------|--------|-----|---------|",
    ]
    for _, row in critical_gaps.iterrows():
        lines.append(
            f"| {row['market']} | {row['persona']} | {row['current_count']} | "
            f"{row['target_count']} | {row['gap_count']} | **{row['urgency_level']}** |"
        )

    lines += [
        "",
        "---",
        "",
        "## 5. What Should I Stop Doing?",
        "",
        "- **Stop connecting with random HR professionals in Bolivia/regional markets** "
        "with no USD or EU relevance. These dilute your network quality.",
        "- **Stop accepting all inbound connection requests** without checking if they "
        "align with your LATAM USD or Spain/EU target.",
        "- **Stop treating data peer connections as strategic** — colleagues and peers "
        "cannot hire you; decision-makers and recruiters can.",
        "- **Stop passive LinkedIn presence** — you need to post content that attracts "
        "LATAM USD and nearshore recruiters to you.",
        "",
        "---",
        "",
        "## 6. What Should I Start Doing?",
        "",
        "1. **Systematic outreach to US/Canada nearshore recruiters** — search for "
        "'Data Engineering recruiter nearshore LATAM' and send 5-10 daily requests.",
        "2. **Engage hiring managers at remote-friendly companies** — Andela, AgileEngine, "
        "Toptal, Crossover, Wizeline, Gorilla Logic, Unosquare.",
        "3. **Post weekly content** on LinkedIn about your Data Engineering expertise "
        "to attract inbound from target recruiters.",
        "4. **Activate your top 50 high-priority contacts** from this report with a "
        "personalized, brief message about your availability.",
        "5. **Begin Spain/EU recruiter network in parallel** — even 5 Spain contacts per "
        "week will compound over 6 months.",
        "",
        "---",
        "",
        "## 7. What Should I Prioritize for the Next 30 Days?",
        "",
        "| Priority | Action | Target | Why |",
        "|----------|--------|--------|-----|",
        "| 1 | Connect with LATAM USD Recruiters | +40 new | Biggest direct USD job pipeline |",
        "| 2 | Connect with US/CA Nearshore Recruiters | +30 new | Critical gap at 72% unfilled |",
        "| 3 | Connect with US/CA Hiring Managers | +20 new | Decision-makers for remote roles |",
        "| 4 | Reactivate top 50 high-priority contacts | 50 messages | Direct activation of existing assets |",
        "| 5 | Post 4x LinkedIn content about Data Engineering | 4 posts | Attract inbound from target recruiters |",
        "",
        "---",
        "_Generated by LinkedIn Connections Heatmap — strategic intelligence layer._",
    ]
    return "\n".join(lines)


def _career_roadmap(kpis: dict) -> str:
    from datetime import date
    today = date.today().strftime("%B %d, %Y")
    return f"""# Career Network Roadmap
> Generated: **{today}**

---

## Context

| | |
|--|--|
| Current location | Brazil |
| Short-term goal | Remote job paid in USD |
| Medium-term goal | Relocate to Spain / Europe |
| Network size | {kpis['total_connections']:,} connections |
| USD Opportunity Score | {kpis['usd_opportunity_score']}/100 |
| Spain Readiness Score | {kpis['spain_readiness_score']}/100 |

---

## 30-Day Roadmap: USD Remote Job Focus (Brazil)

### Goal
Land or advance toward a remote Data Engineering contract paid in USD.

### Actions
| Week | Focus | Daily Target | Weekly Goal |
|------|-------|-------------|-------------|
| Week 1 | LATAM USD Recruiters (AgileEngine, Andela, LatamCent, Wizeline) | 5 connections/day | 35 new connections |
| Week 2 | US/Canada Nearshore Recruiters | 5 connections/day | 35 new connections |
| Week 3 | US/Canada & LATAM Hiring Managers | 4 connections/day | 28 new connections |
| Week 4 | Activate existing top-priority contacts | 10 DMs/day | 70 messages sent |

### Key LinkedIn Search Strategies
- `"data engineer" "nearshore" "LATAM" recruiter`
- `"data engineering manager" "remote" "contractor"`
- `site:linkedin.com "hiring data engineer" "remote" "LATAM"`

### Saturated Areas (No Action Needed)
- Brazil generic recruiters (41 existing — target already met)
- LATAM USD Directors (29 existing — target met)

---

## 60-Day Roadmap: Extend USD + Begin Spain/EU

### USD Maintenance (Days 31-60)
- Continue adding 2-3 US/Canada hiring managers per week
- Engage existing high-priority USD contacts with content reactions
- Track response rates from 30-day outreach

### Spain/EU Launch (Days 31-60)
| Week | Focus | Daily Target |
|------|-------|-------------|
| Week 5-6 | Spain/EU Recruiters (Madrid, Barcelona, Dublin) | 3 connections/day |
| Week 7-8 | Spain/EU Talent Acquisition | 3 connections/day |

### Key Spain/EU Companies to Target
- **Spain**: Capgemini Spain, NTT Data Spain, ERNI, Seidor, Stratesys, Indra, Minsait
- **Portugal**: Volkswagen Group Portugal, DataSmart, emagine Portugal
- **Ireland**: Accenture Dublin, Google Dublin, Amazon AWS Dublin
- **Germany**: Adesso SE, SAP, Deutsche Telekom, Zalando

---

## 90-Day Roadmap: Balance USD Income + EU Positioning

### USD Stability (Days 61-90)
- Maintain recruiter relationships with monthly check-ins
- Target: 1 active USD job conversation by Day 90
- Focus on Data Engineering Manager connections in US/CA nearshore

### EU Deep Build (Days 61-90)
| Week | Focus |
|------|-------|
| Week 9-10 | Spain/EU Hiring Managers + Heads of Data |
| Week 11-12 | European Data Leaders + referral contacts (Germany, Netherlands) |

### Success Metrics at Day 90
| KPI | Target |
|-----|--------|
| LATAM USD Recruiters | 80 |
| US/CA Nearshore Recruiters | 50 |
| USD Hiring Managers | 30 |
| Spain/EU Recruiters | 20 |
| Spain/EU Hiring Managers | 15 |
| USD Opportunity Score | >= 60 |
| Spain Readiness Score | >= 35 |

---

## Priority Market This Week
**{kpis['top_priority_market']}** — This is where your biggest immediate gap is vs. your short-term goal.

## Priority Market Before Moving to Spain
**SPAIN_EU + EUROPE** — You need at minimum 20 Spain/EU recruiters and 15 hiring managers
before relocating. Start this effort at Day 30 and sustain through Day 90.

---

_Generated by LinkedIn Connections Heatmap — career intelligence layer._
"""


def _kpi_report(kpis: dict) -> str:
    from datetime import date
    today = date.today().strftime("%B %d, %Y")
    return f"""# KPI Dashboard Report
> Generated: **{today}**

This document explains every KPI in the dashboard.

---

## Core Counts

| KPI | Value | Definition |
|-----|-------|-----------|
| Total Connections | {kpis['total_connections']:,} | All valid connections in the export |
| High Priority | {kpis['high_priority']:,} | Priority score >= 70 |
| Medium Priority | {kpis['medium_priority']:,} | Priority score 40-69 |
| Low Priority | {kpis['low_priority']:,} | Priority score < 40 |

---

## Persona Groups

| KPI | Value | Personas Included |
|-----|-------|------------------|
| Recruiters | {kpis['recruiters_total']:,} | Recruiter, Sourcer |
| Talent/HR | {kpis['talent_hr_total']:,} | Talent Acquisition, HR |
| Hiring Managers | {kpis['hiring_managers_total']:,} | Hiring Manager, Engineering Manager |
| Data Leaders | {kpis['data_leaders_total']:,} | Data Engineering Manager, Head of Data, Director, Executive |
| Data Peers | {kpis['data_peers_total']:,} | Data Engineer, Analytics Engineer, Data Analyst, BI Analyst, Data Scientist, ML/AI |

---

## Market Counts

| Market | Count | Description |
|--------|-------|-------------|
| BRAZIL | {kpis['brazil_count']:,} | Connections with Brazil market signals |
| LATAM_USD | {kpis['latam_usd_count']:,} | LATAM companies paying USD / nearshore platforms |
| US_CANADA_NEARSHORE | {kpis['us_nearshore_count']:,} | US/Canada companies with LATAM contractor culture |
| SPAIN_EU | {kpis['spain_eu_count']:,} | Spain and Portugal specifically |
| EUROPE | {kpis['europe_count']:,} | Other European markets (Germany, Netherlands, Ireland, etc.) |
| UNKNOWN | {kpis['unknown_count']:,} ({kpis['unknown_pct']}%) | No market keyword matched — not necessarily irrelevant |

---

## Strategic Scores

### USD Opportunity Score: {kpis['usd_opportunity_score']}/100
Weighted composite measuring readiness to land a USD remote job from Brazil.
- Recruiters in USD markets (30 pts): {kpis['usd_recruiters']} / target 60
- TA in USD markets (20 pts): {kpis['usd_ta']} / target 40
- Hiring Mgrs in USD markets (20 pts): {kpis['usd_hiring_mgrs']} / target 30
- Data Leaders in USD markets (15 pts): {kpis['usd_data_leaders']} / target 20
- High-priority USD contacts (15 pts): {kpis['usd_high_priority']} / target 30

### Spain Readiness Score: {kpis['spain_readiness_score']}/100
Measures how prepared your network is for a Spain/Europe move.
- Spain/EU recruiters (30 pts): {kpis['spain_recruiters']} / target 40
- Spain/EU TA (20 pts): {kpis['spain_ta']} / target 30
- Spain/EU Hiring Mgrs (20 pts): {kpis['spain_hiring_mgrs']} / target 20
- EU Data Leaders (15 pts): {kpis['spain_data_leaders']} / target 15
- High-priority Spain/EU contacts (15 pts): {kpis['spain_high_priority']} / target 15

### Market Readiness Score: {kpis['market_readiness_score']}/100
Composite of USD (60%) + Spain (40%) scores.

---

## Priority Score Calculation

Each connection receives a score 0-100 based on:
| Factor | Max Points | Description |
|--------|-----------|-------------|
| Persona relevance | 35 | Recruiter/TA = high, Other = low |
| Strategic market | 30 | US_CANADA_NEARSHORE = max, UNKNOWN = min |
| Seniority | 15 | Executive/Founder = high, Intern = low |
| Recency | 10 | Connected in last 30 days = full bonus |
| Company signal | 10 | Tech/data/nearshore company keywords |

---

## Data Quality Flags

- Missing company: {kpis['missing_company_count']:,} ({kpis['missing_company_pct']}%)
- Missing position: {kpis['missing_position_count']:,} ({kpis['missing_position_pct']}%)
- Unknown market: {kpis['unknown_count']:,} ({kpis['unknown_pct']}%)

**Why is unknown market so high?**
LinkedIn does not include location in its bulk connection export. Market inference
uses only company name and job title keywords. This is a fundamental LinkedIn API
limitation — not a bug in the analysis.

---

_Generated by LinkedIn Connections Heatmap — KPI intelligence layer._
"""


# ── Excel update ─────────────────────────────────────────────────────────────

def update_excel_dashboard(
    kpis: dict, plan_30: pd.DataFrame, plan_60: pd.DataFrame,
    plan_90: pd.DataFrame, market_matrix: pd.DataFrame,
    persona_matrix: pd.DataFrame, gap_matrix: pd.DataFrame,
) -> None:
    """Append new strategy sheets to the existing Excel workbook."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Load existing workbook
        wb = load_workbook(DASHBOARD_XLSX)

        # Build KPI DataFrame
        kpi_rows = [{"KPI": k, "Value": str(v)} for k, v in kpis.items()
                    if not isinstance(v, (dict, list))]
        kpi_df = pd.DataFrame(kpi_rows)

        # Sheets to add
        new_sheets = {
            "Executive KPI Summary":    kpi_df,
            "Market Strategy Matrix":   market_matrix,
            "Persona Strategy Matrix":  persona_matrix,
            "Gap Matrix":               gap_matrix,
            "30 Day Action Plan":       plan_30,
            "60 Day Action Plan":       plan_60,
            "90 Day Action Plan":       plan_90,
        }

        header_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name, df in new_sheets.items():
            # Remove if exists
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(title=sheet_name)

            # Write header
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Write data
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-fit columns
            for col_cells in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in col_cells if c.value), default=10
                )
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

        # Add KPI dictionary sheet
        if "KPI Dictionary" in wb.sheetnames:
            del wb["KPI Dictionary"]
        ws_dict = wb.create_sheet("KPI Dictionary")
        dict_data = [
            ("KPI", "Definition", "Formula"),
            ("total_connections", "Total valid connections analyzed", "count(classified_connections)"),
            ("high_priority", "Connections with score >= 70", "sum(score >= 70)"),
            ("usd_opportunity_score", "Readiness for USD remote job (0-100)", "weighted(recruiters+ta+hm+dl+high_priority in USD markets)"),
            ("spain_readiness_score", "Readiness for Spain/EU move (0-100)", "weighted(recruiters+ta+hm+dl+high_priority in SPAIN+EU markets)"),
            ("market_readiness_score", "Composite score (0-100)", "usd_score*0.6 + spain_score*0.4"),
            ("unknown_pct", "% of connections without market inference", "unknown_count / total * 100"),
        ]
        for row in dict_data:
            ws_dict.append(row)

        wb.save(DASHBOARD_XLSX)
        logger.info(f"Excel dashboard updated with new sheets: {DASHBOARD_XLSX.name}")

    except Exception as e:
        logger.warning(f"Could not update Excel: {e}")


# ── Main pipeline ─────────────────────────────────────────────────────────────

def run_strategy_layer() -> None:
    t0 = time.time()
    logger.info("=" * 60)
    logger.info("  Strategy Layer — Starting")
    logger.info("=" * 60)

    # 1. Load classified data
    logger.info("Step 1/7: Loading classified_connections.csv …")
    if not CLASSIFIED_CSV.exists():
        logger.error(f"Not found: {CLASSIFIED_CSV}")
        logger.error("Run `python src/build_network_heatmap.py` first.")
        sys.exit(1)

    df = pd.read_csv(CLASSIFIED_CSV, dtype=str, low_memory=False)
    df["priority_score"]    = pd.to_numeric(df["priority_score"],    errors="coerce").fillna(0)
    df["market_confidence"] = pd.to_numeric(df["market_confidence"], errors="coerce").fillna(0)
    logger.info(f"  Loaded {len(df):,} rows")

    # 2. KPIs
    logger.info("Step 2/7: Computing KPIs …")
    kpis = compute_kpis(df)
    _save_kpi_csv(kpis)
    logger.info(f"  USD Opportunity Score:  {kpis['usd_opportunity_score']}")
    logger.info(f"  Spain Readiness Score:  {kpis['spain_readiness_score']}")

    # 3. Action plans
    logger.info("Step 3/7: Building action plans …")
    plan_30 = build_30_day_plan(df)
    plan_60 = build_60_day_plan(df)
    plan_90 = build_90_day_plan(df)
    _save_csv(plan_30, PLAN_30, "30-Day Plan")
    _save_csv(plan_60, PLAN_60, "60-Day Plan")
    _save_csv(plan_90, PLAN_90, "90-Day Plan")

    # 4. Strategy matrices
    logger.info("Step 4/7: Building strategy matrices …")
    market_matrix  = build_market_strategy_matrix(df)
    persona_matrix = build_persona_strategy_matrix(df)
    gap_matrix     = build_connection_gap_matrix(df)
    _save_csv(market_matrix,  MARKET_MATRIX,  "Market Strategy Matrix")
    _save_csv(persona_matrix, PERSONA_MATRIX, "Persona Strategy Matrix")
    _save_csv(gap_matrix,     GAP_MATRIX,     "Connection Gap Matrix")

    # 5. Dashboard data
    logger.info("Step 5/7: Building dashboard data …")
    company_intel = build_company_intelligence(df)
    heatmaps_data = build_heatmap_data(df)
    top_contacts  = build_top_contacts(df, n=100)
    save_dashboard_json(kpis, heatmaps_data, company_intel, top_contacts,
                        plan_30, plan_60, plan_90, gap_matrix)

    # 6. Markdown reports
    logger.info("Step 6/7: Generating Markdown reports …")
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    exec_report = _executive_report(kpis, gap_matrix)
    EXEC_REPORT.write_text(exec_report, encoding="utf-8")
    logger.info(f"  Saved: {EXEC_REPORT.name}")

    career_road = _career_roadmap(kpis)
    CAREER_ROADMAP.write_text(career_road, encoding="utf-8")
    logger.info(f"  Saved: {CAREER_ROADMAP.name}")

    kpi_rep = _kpi_report(kpis)
    KPI_REPORT.write_text(kpi_rep, encoding="utf-8")
    logger.info(f"  Saved: {KPI_REPORT.name}")

    # 7. Update Excel
    logger.info("Step 7/7: Updating Excel dashboard …")
    if DASHBOARD_XLSX.exists():
        update_excel_dashboard(kpis, plan_30, plan_60, plan_90,
                               market_matrix, persona_matrix, gap_matrix)
    else:
        logger.warning("dashboard_ready.xlsx not found — skipping Excel update.")

    elapsed = round(time.time() - t0, 1)
    logger.info("=" * 60)
    logger.info(f"  Strategy layer complete in {elapsed}s")
    logger.info("=" * 60)
    logger.info("[New Outputs]")
    logger.info("  outputs/kpi_summary.csv")
    logger.info("  outputs/action_plan_30_days.csv")
    logger.info("  outputs/action_plan_60_days.csv")
    logger.info("  outputs/action_plan_90_days.csv")
    logger.info("  outputs/connection_gap_matrix.csv")
    logger.info("  outputs/market_strategy_matrix.csv")
    logger.info("  outputs/persona_strategy_matrix.csv")
    logger.info("  outputs/dashboard_metrics.json")
    logger.info("  reports/executive_strategy_report.md")
    logger.info("  reports/career_network_roadmap.md")
    logger.info("  reports/kpi_dashboard_report.md")
    logger.info("  reports/dashboard_ready.xlsx (updated)")
    logger.info("")
    logger.info("  Run dashboard: streamlit run app/dashboard.py")


if __name__ == "__main__":
    try:
        run_strategy_layer()
    except Exception as e:
        logger.exception(f"Strategy layer failed: {e}")
        sys.exit(1)
